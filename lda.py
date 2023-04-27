import numpy as np
import os
import jieba
import random
from gensim.models.ldamodel import LdaModel
from gensim.corpora.dictionary import Dictionary
from collections import Counter
import lda
from sklearn.cluster import KMeans
import openpyxl

with open('./stopwords.txt', 'r', encoding='utf-8') as f:
    stopwords = set([line.strip() for line in f])
# 读取小说语料库中的所有文本
corpus_path = 'txt'
texts = []
for file in os.listdir(corpus_path):
    with open(os.path.join(corpus_path, file), 'r', encoding='ansi') as f:
        text = f.read()
        # 将每个文本按句子分割成段落
        paragraphs = text.split('\n')
        for p in paragraphs:
            if len(p) > 500:  # 只选取长度大于500的段落
                texts.append((p, file[:-4]))  # 每个段落的标签即为所属小说的文件名（去除后缀）

# 随机选择200个段落
random.shuffle(texts)
texts = texts[:200]
texts_chars = []
topic_chars = []
for text, topic in texts:
    words = list(word for word in jieba.cut(text) if word not in stopwords)
    wod = []
    for word in words:
        if '\u4e00' <= word <= '\u9fa5':
            wod.append(word)
    wod = [char for word in wod for char in word]  # 将每个词转化为其包含的字
    texts_chars.append(wod)
    topic_chars.append(topic)

corpus = texts_chars
# 构建词典
vocab = set(word for doc in corpus for word in doc)
word2id = dict((v, idx) for idx, v in enumerate(vocab))
# 将文本数据转换为词频矩阵
M = len(corpus)
V = len(vocab)
X = np.zeros((M, V))
for i, doc in enumerate(corpus):
    for word in doc:
        X[i][word2id[word]] += 1
num_topics_ = len(set(topic_chars))
# 训练LDA模型
model = lda.LDA(n_topics=num_topics_, n_iter=2000, random_state=1)
X = X.astype(int)
model.fit(X)
# 输出每个主题的单词分布
topic_word = model.topic_word_
for i, topic_dist in enumerate(topic_word):
    topic_words = np.array(list(vocab))[np.argsort(topic_dist)][:-(10+1):-1]
    print('主题%d的单词分布：' % i, topic_words)
# 获得训练文本所属主题信息
doc_topic = model.transform(X)
for i in range(len(corpus)):
    print('文本%d的主题分布：' % i, doc_topic[i])
# 定义KMeans聚类模型，假设要将文本集合聚为n类
kmeans = KMeans(n_clusters=num_topics_, random_state=0)
# 训练KMeans模型
kmeans.fit(doc_topic)
# 获得每个文本所属的簇
labels = kmeans.labels_
for i in range(len(labels)):
    print('文本%d所属簇为：%d' % (i, labels[i]))
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = 'Sheet1'
worksheet['A1'] = 'Name'
worksheet['B1'] = '类别'
i = 0
for topic in topic_chars:
    row = i + 2
    worksheet.cell(row=row, column=1, value=topic)
    worksheet.cell(row=row, column=2, value=labels[i])
    i += 1
workbook.save('data1.xlsx')